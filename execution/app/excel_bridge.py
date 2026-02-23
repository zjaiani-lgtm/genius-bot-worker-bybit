from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook  # type: ignore

from .logger import log
from .utils import safe_float


@dataclass
class ExcelDecision:
    decision: str  # BUY / SELL / NO
    confidence: float
    meta: Dict[str, Any]


class ExcelBridge:
    """
    Headless Excel bridge using openpyxl.

    IMPORTANT:
    openpyxl does NOT calculate Excel formulas.
    It can read cached formula results if the file was saved with results.
    """

    def __init__(self, path: str, sheet: str, in_prefix: str, out_prefix: str, logger):
        self.path = path
        self.sheet = sheet
        self.in_prefix = in_prefix
        self.out_prefix = out_prefix
        self.logger = logger

    def _resolve_named_cell(self, wb, name: str) -> Optional[Tuple[str, str]]:
        dn = wb.defined_names.get(name)
        if dn is None:
            return None
        dests = list(dn.destinations)
        if not dests:
            return None
        sheet, coord = dests[0]
        return sheet, coord

    def _set_named(self, wb, name: str, value: Any) -> bool:
        resolved = self._resolve_named_cell(wb, name)
        if not resolved:
            return False
        sheet, coord = resolved
        wb[sheet][coord].value = value
        return True

    def _get_named(self, wb, name: str) -> Any:
        resolved = self._resolve_named_cell(wb, name)
        if not resolved:
            return None
        sheet, coord = resolved
        return wb[sheet][coord].value

    def evaluate(self, symbol: str, timeframe: str, last: float, atr_pct: float, trend_score: float, vol_score: float) -> ExcelDecision:
        if not self.path or not os.path.exists(self.path):
            return ExcelDecision(decision="NO", confidence=0.0, meta={"error": "excel_path_missing"})

        try:
            wb = load_workbook(self.path, data_only=True)

            self._set_named(wb, f"{self.in_prefix}SYMBOL", symbol)
            self._set_named(wb, f"{self.in_prefix}TF", timeframe)
            self._set_named(wb, f"{self.in_prefix}LAST", last)
            self._set_named(wb, f"{self.in_prefix}ATR_PCT", atr_pct)
            self._set_named(wb, f"{self.in_prefix}TREND", trend_score)
            self._set_named(wb, f"{self.in_prefix}VOL", vol_score)

            decision = self._get_named(wb, f"{self.out_prefix}DECISION")
            conf = self._get_named(wb, f"{self.out_prefix}CONF")

            if decision is None:
                ws = wb[self.sheet] if self.sheet in wb.sheetnames else wb.active
                decision = ws[os.getenv("EXCEL_DECISION_CELL", "B2")].value

            if conf is None:
                ws = wb[self.sheet] if self.sheet in wb.sheetnames else wb.active
                conf = ws[os.getenv("EXCEL_CONF_CELL", "B3")].value

            decision_s = str(decision).strip().upper() if decision is not None else "NO"
            if decision_s not in ("BUY", "SELL", "NO"):
                decision_s = "NO"

            return ExcelDecision(decision=decision_s, confidence=safe_float(conf, 0.0), meta={})
        except Exception as e:
            log(self.logger, "WARNING", "EXCEL_EVAL_FAIL", error=str(e))
            return ExcelDecision(decision="NO", confidence=0.0, meta={"error": str(e)})
